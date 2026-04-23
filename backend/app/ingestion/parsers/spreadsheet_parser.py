import csv
import io
from app.ingestion.chunker import RawChunk, chunk_markdown


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_xlsx(file_bytes: bytes, doc_title: str) -> list[RawChunk]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        table_lines: list[str] = []
        header: list[str] | None = None

        for row in ws.iter_rows(values_only=True):
            cells = [_cell_str(c) for c in row]
            if not any(cells):
                continue
            if header is None:
                header = cells
                table_lines.append("| " + " | ".join(header) + " |")
                table_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            else:
                # Pad or trim row to match header width
                while len(cells) < len(header):
                    cells.append("")
                table_lines.append("| " + " | ".join(cells[: len(header)]) + " |")

        if table_lines:
            sections.append(f"## {sheet_name}\n\n" + "\n".join(table_lines))

    wb.close()
    content = "\n\n".join(sections)
    return chunk_markdown(content, source_title=doc_title, source_path=doc_title)


def parse_csv(file_bytes: bytes, doc_title: str) -> list[RawChunk]:
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader if any(c.strip() for c in row)]

    if not rows:
        return []

    header = rows[0]
    table_lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * len(header)) + " |",
    ]
    for row in rows[1:]:
        while len(row) < len(header):
            row.append("")
        table_lines.append("| " + " | ".join(row[: len(header)]) + " |")

    content = "\n".join(table_lines)
    return chunk_markdown(content, source_title=doc_title, source_path=doc_title)
